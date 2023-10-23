###### project 1 automation #### 
# python programme that can process a thousand spreadsheet under a second.
# 
## how to process spreadsheet using python, this is valuable where you have hundred or thousand of spreadsheet
# that need to be updated.

#  due to an error, the values in the price coulumn in the excel spreadsheet "transaction_python_project_1",
# was wrong, and we need to correct those values,
# so assume we need to decrease the each value in the "price column" by 10 %, then manually, we would do,
# in excel create another column, then add a formula like "=C2*0.9", 
# where C2 is the cell in excel containg a value under "price column", so we need
# to do this for each row in the "price column", but what about if the 
# the price column had 5,000 rows , then we would need to scroll and up and down, and thier is chance of error.

# python can be used to fix this, also python can be used to add chart to alot of spread sheet very quickly,
### 


import openpyxl as xl 
from openpyxl.chart import BarChart, Reference #here BarChart, Reference  are "classes" because they start with cappital letter, also name of functions 
# if we need to update thousand excel file, then this program wouul be difficult to use,
# because we need to write the excel file name each time, this is time consuming,
# but what we can do is create a function , that has an inpur of an excel file name
# and the function does the operations below:
def process_workbook(filename):
    wb=xl.load_workbook(filename)
    sheet=wb["Sheet1"]
# getting cells in 3rd column
# we do not want the 1st row because it contains heading,
    for row in range(2, sheet.max_row+1): # adding 1 to sheet.max_row because range function does not include last value
        cell=sheet.cell(row,3) # here since the "price" column is the 3rd column in trasaction excel file, then there is 3
        corrected_price=cell.value*0.9
     # here cell.value is the value of the cell
        corrected_price_cell=sheet.cell(row,4)
        corrected_price_cell.value=corrected_price

# here the argument in "save" function (and wb is an object containing save function')
# is the name of the file we want to save it as,

## add new column on our column on our work sheet, that containes the
# corrected price value

### how to add a chart to the excel file ""
# "from openpyxl.chart import BarChart, Reference"
# then select a range of values in excel sheet
# we want to select all values (from row 2 to 4) in 4th column in "project 1" excel file in sheet called "Sheet1"
# use the "Reference" class in "Chart" file to select range of values in a sheet in excel file (below "values" is an object of class "Reference")
    values=Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4,max_col=4) # here we are selecting all cells from and including row 2 to row n, where n is the man number of filled in rows of the sheet, but we want the cells to be from the 4th column (not from other columns)
# here the "value" object of class "Reference", selects values in a range of cells  in the excel file
# now going to create an object of Barchart class called "chart":
    chart=BarChart()
# then when we put in "values" into the function called "add_data" in "chart" object,  an output of a chart
    chart.add_data(values) # here "add_data" is a function in object "chart"
# now we need to add the chart into the sheet called "Sheet1" in Excel file:
    sheet.add_chart(chart,'g2') # here the 2nd argument of the function "add_chart" is location in excel of where 
# we want to the chart to be placed, so "g2" is the cell location in excel file, where we want the chart to be located in.
#from testfolder import test1 as t1
#print(t1.add(0,3))
# below we need to save the updated excel file "transaction_python_project_1.xlsx"
# as "transaction_updated.xlsx" by using the following code:
    wb.save(filename)

    # automation, 
